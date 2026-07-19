"""Extract French Yu-Gi-Oh! card names from a personal collection workbook."""
from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

from openpyxl import load_workbook


EXCLUDED_LANGUAGES = {'English', 'Portuguese', 'German', 'Spanish', 'Italian'}
CACHE_PATH = Path(__file__).resolve().parents[1] / 'cache' / 'card_names_fr_collection.json'


def is_french_collection_row(language: object) -> bool:
    """Blank language cells are French by default in this collection workbook."""
    if language is None or (isinstance(language, float) and math.isnan(language)):
        return True
    return str(language).strip() not in EXCLUDED_LANGUAGES


def normalise_name(value: object) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ''
    return str(value).strip()


def extract_names(workbook_path: Path) -> tuple[int, int, int, list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if 'Liste' not in workbook.sheetnames:
            raise ValueError('La feuille « Liste » est introuvable.')
        sheet = workbook['Liste']
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError('La feuille « Liste » est vide.')
        columns = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
        required = {'Nom de la carte', 'Langue'}
        missing = required - columns.keys()
        if missing:
            raise ValueError(f'Colonnes manquantes : {", ".join(sorted(missing))}.')

        rows_read = 0
        raw_names: list[str] = []
        for row in rows:
            rows_read += 1
            language = row[columns['Langue']] if columns['Langue'] < len(row) else None
            if not is_french_collection_row(language):
                continue
            name = normalise_name(row[columns['Nom de la carte']] if columns['Nom de la carte'] < len(row) else None)
            if name:
                raw_names.append(name)
        names = sorted(set(raw_names), key=str.casefold)
        return rows_read, len(raw_names), len(raw_names) - len(names), names
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description='Extrait les noms français depuis une collection Excel YGO.')
    parser.add_argument('workbook', type=Path, help='Chemin du fichier Excel .xlsx')
    args = parser.parse_args()
    if not args.workbook.is_file():
        parser.error(f'Fichier introuvable : {args.workbook}')

    rows_read, raw_name_count, duplicates_removed, names = extract_names(args.workbook)
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(names, ensure_ascii=False), encoding='utf-8')
    print(f'Lignes lues : {rows_read}')
    print(f'Noms français extraits : {len(names)}')
    print(f'Doublons retirés : {duplicates_removed}')
    print(f'Noms source avant dédoublonnage : {raw_name_count}')
    print(f'Cache écrit : {CACHE_PATH}')


if __name__ == '__main__':
    main()
